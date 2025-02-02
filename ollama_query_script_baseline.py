import openpyxl
from langchain_ollama.llms import OllamaLLM
from tqdm import tqdm


# Loads prompts from the first column of an Excel file
def load_prompts(file_path):

    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    prompts = [sheet.cell(row=i, column=1).value for i in range(2, sheet.max_row + 1) if sheet.cell(row=i, column=1).value]
    return prompts


# Queries the Ollama model with a given prompt
def query_ollama_model(model, prompt):

    model = OllamaLLM(model=model)
    response = model.invoke(prompt)
    return response


# Saves responses to the first column of a new Excel file
def save_responses(file_path, responses):

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for i, response in enumerate(responses, start=1):
        sheet.cell(row=i, column=1, value=response)
    workbook.save(file_path)


if __name__ == "__main__":

    # Modify the model name and file paths as needed
    model_name = "llama2"
    input_file_path = "strategyqa_prompts_llama_phi.xlsx"
    output_file_path = "strategyqa_llama_responses.xlsx"

    # Load prompts from the input Excel file
    prompts = load_prompts(input_file_path)
    responses = []

    # Process each prompt in the file
    for prompt in tqdm(prompts, desc="Processing Prompts", unit="prompt"):
        response = query_ollama_model(model_name, prompt)
        responses.append(response)

    # Save all responses to the output Excel file
    save_responses(output_file_path, responses)